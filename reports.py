"""
Excel report generator for MAA Payment Record Management System.
Produces in-memory .xlsx bytes using openpyxl.
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill as StylePatternFill
from openpyxl.utils import get_column_letter

# ── Colours ───────────────────────────────────────────────────────────────────
HEADER_FILL = StylePatternFill(fill_type="solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri")

DOCTOR_COPY_HEADER_FILL = StylePatternFill(fill_type="solid", fgColor="00695C")
DOCTOR_COPY_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri")

GREEN_FILL  = StylePatternFill(fill_type="solid", fgColor="C6EFCE")
AMBER_FILL  = StylePatternFill(fill_type="solid", fgColor="FFEB9C")
TOTAL_FILL  = StylePatternFill(fill_type="solid", fgColor="D9E1F2")
TOTAL_FONT  = Font(bold=True, name="Calibri")

ROW_FILL_ODD  = StylePatternFill(fill_type="solid", fgColor="FFFFFF")
ROW_FILL_EVEN = StylePatternFill(fill_type="solid", fgColor="EEF2F9")

RUPEE_FMT   = '₹#,##0.00'
NUMBER_FMT  = '#,##0'
DATE_FMT    = 'dd-mmm-yyyy'

# ── Column definitions per report type ───────────────────────────────────────

ADMISSION_COLS = [
    ("TID",             "tid",             None),
    ("Patient Name",    "patient_name",    None),
    ("Gender",          "gender",          None),
    ("Age",             "age",             NUMBER_FMT),
    ("Admission Date",  "date_of_admission", DATE_FMT),
    ("Discharge Date",  "date_of_discharge", DATE_FMT),
    ("Days",            "days",            NUMBER_FMT),
    ("# Packages",      "packages",        NUMBER_FMT),
    ("Total Approved",  "total_approved",  RUPEE_FMT),
    ("Total Paid",      "total_paid",      RUPEE_FMT),
    ("Received (−TDS)", "total_received",  RUPEE_FMT),
    ("Outstanding",     "outstanding",     RUPEE_FMT),
    ("Query Count",     "queries",         NUMBER_FMT),
    ("Statuses",        "statuses",        None),
]

MONTHLY_COLS = [
    ("Month",           "month",           None),
    ("Admissions",      "admissions",      NUMBER_FMT),
    ("Packages",        "packages",        NUMBER_FMT),
    ("Total Approved",  "total_approved",  RUPEE_FMT),
    ("Total Paid",      "total_paid",      RUPEE_FMT),
    ("Received (−TDS)", "total_received",  RUPEE_FMT),
    ("Outstanding",     "outstanding",     RUPEE_FMT),
]

FY_COLS = [
    ("Financial Year",  "financial_year",  None),
    ("Admissions",      "admissions",      NUMBER_FMT),
    ("Packages",        "packages",        NUMBER_FMT),
    ("Total Approved",  "total_approved",  RUPEE_FMT),
    ("Total Paid",      "total_paid",      RUPEE_FMT),
    ("Received (−TDS)", "total_received",  RUPEE_FMT),
    ("Outstanding",     "outstanding",     RUPEE_FMT),
]

AMOUNT_COLS = {
    "total_approved", "total_paid", "total_received", "outstanding",
    "approved_amount", "paid_amount", "pkg_rate",
    "hosp_ex", "pharma_ex", "dialysis_ex", "total_ex",
    "maa_payment", "doctor_share", "hospital_share",
}
SUMMABLE_COLS = {
    "total_approved", "total_paid", "total_received", "outstanding",
    "approved_amount", "paid_amount", "packages",
    "admissions", "queries", "query_raised", "days",
    "hosp_ex", "pharma_ex", "dialysis_ex", "total_ex",
    "maa_payment", "doctor_share", "hospital_share",
}

DOCTOR_INTERNAL_COLS = [
    ("No.",               "no",                   NUMBER_FMT),
    ("Patient Name",      "patient_name",         None),
    ("Admission Date",    "admission_date",        DATE_FMT),
    ("Hosp Ex",           "hosp_ex",              RUPEE_FMT),
    ("Pharma Ex",         "pharma_ex",            RUPEE_FMT),
    ("Dialysis Ex",       "dialysis_ex",          RUPEE_FMT),
    ("Total Ex",          "total_ex",             RUPEE_FMT),
    ("MAA Payment",       "maa_payment",          RUPEE_FMT),
    ("Doctor Share",      "doctor_share",         RUPEE_FMT),
    ("Hospital Share",    "hospital_share",       RUPEE_FMT),
    ("MAA Status",        "maa_status",           None),
    ("Doctor Paid",       "doctor_paid_label",    None),
    ("Dr Payment Month",  "doctor_payment_month", None),
    ("Comments",          "comments",             None),
]

DOCTOR_COPY_COLS = [
    ("No.",               "no",                   NUMBER_FMT),
    ("Patient Name",      "patient_name",         None),
    ("Admission Date",    "admission_date",        DATE_FMT),
    ("Hosp Ex",           "hosp_ex",              RUPEE_FMT),
    ("Pharma Ex",         "pharma_ex",            RUPEE_FMT),
    ("Dialysis Ex",       "dialysis_ex",          RUPEE_FMT),
    ("Total Ex",          "total_ex",             RUPEE_FMT),
    ("MAA Payment",       "maa_payment",          RUPEE_FMT),
    ("Doctor Share",      "doctor_share",         RUPEE_FMT),
    ("Hospital Share",    "hospital_share",       RUPEE_FMT),
    ("Comments",          "comments",             None),
]


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if isinstance(cell.value, (int, float)) and cell.number_format not in (None, "", "General"):
                    # Measure the formatted display width, not the raw float string.
                    # e.g. 45678.0 → "45,678.00" + 2 for currency symbol = 11
                    length = len(f"{cell.value:,.2f}") + 2
                else:
                    length = len(str(cell.value or ""))
                max_len = max(max_len, length)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def _write_sheet(ws, df: pd.DataFrame, col_defs: list[tuple], title: str,
                 status_col_idx: int | None = None,
                 color_by_maa_status: bool = False,
                 header_fill=None, header_font=None):
    """Write headers + data + summary row into ws."""
    ws.title = title[:31]  # Excel sheet name limit

    _hfill = header_fill if header_fill is not None else HEADER_FILL
    _hfont = header_font if header_font is not None else HEADER_FONT

    headers = [c[0] for c in col_defs]
    db_keys = [c[1] for c in col_defs]
    fmts    = [c[2] for c in col_defs]

    # Header row
    for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill = _hfill
        cell.font = _hfont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, (key, fmt) in enumerate(zip(db_keys, fmts), 1):
            val = row.get(key)
            if not isinstance(val, str) and pd.isna(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            if fmt:
                cell.number_format = fmt

        # Row fill: status-based overrides alternating
        alt_fill = ROW_FILL_EVEN if ri % 2 == 0 else ROW_FILL_ODD
        if color_by_maa_status:
            maa_status = str(row.get("maa_status", "") or "")
            if maa_status == "Claim Paid":
                row_fill = GREEN_FILL
            elif maa_status == "Rejected":
                row_fill = _REJECTED_FILL
            else:
                row_fill = alt_fill
        elif status_col_idx is not None:
            statuses_val = str(row.get("statuses", "") or "")
            statuses = {s.strip().lower() for s in statuses_val.split(",")}
            queries  = row.get("queries", 0) or 0
            if queries > 0:
                row_fill = AMBER_FILL
            elif statuses and all("paid" in s for s in statuses if s):
                row_fill = GREEN_FILL
            else:
                row_fill = alt_fill
        else:
            row_fill = alt_fill

        for ci in range(1, len(col_defs) + 1):
            ws.cell(row=ri, column=ci).fill = row_fill

    # Summary row
    last_data_row = len(df) + 1
    summary_row   = last_data_row + 1
    ws.cell(row=summary_row, column=1, value="TOTAL").font = TOTAL_FONT

    for ci, (key, fmt) in enumerate(zip(db_keys, fmts), 1):
        if key in SUMMABLE_COLS and pd.api.types.is_numeric_dtype(df[key]) if key in df.columns else False:
            total = df[key].sum()
            cell  = ws.cell(row=summary_row, column=ci, value=total)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            if fmt:
                cell.number_format = fmt
        else:
            ws.cell(row=summary_row, column=ci).fill = TOTAL_FILL

    _auto_width(ws)


_MONTH_NAMES = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May",     "06": "June",     "07": "July",  "08": "August",
    "09": "September","10": "October", "11": "November","12": "December",
}

_DETAIL_HEADERS = [
    "TID", "Patient Name", "Admission Date", "Discharge Date",
    "Paid", "Received (−TDS)", "Approved (Not Paid)", "Rejected",
]
_DETAIL_KEYS = [
    "tid", "patient_name", "date_of_admission", "date_of_discharge",
    "paid", "received", "approved", "rejected",
]
_DETAIL_FMTS = [None, None, DATE_FMT, DATE_FMT, RUPEE_FMT, RUPEE_FMT, RUPEE_FMT, RUPEE_FMT]

SUBTOTAL_FILL = StylePatternFill(fill_type="solid", fgColor="D9E1F2")
SUBTOTAL_FONT = Font(bold=True, name="Calibri")


def month_label(ym: str) -> str:
    """'2025-04' → 'April 2025'"""
    try:
        year, month = ym.split("-")
        return f"{_MONTH_NAMES.get(month, month)} {year}"
    except Exception:
        return ym


def _generate_detail_report(df: pd.DataFrame, sheet_title: str) -> bytes:
    """
    Shared logic for FY and Month admission detail reports.
    Layout: column header row → per-month sections (header, data, subtotal, blank) → grand total.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    num_cols = len(_DETAIL_HEADERS)

    # Row 1: column header
    for ci, hdr in enumerate(_DETAIL_HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    current_row = 2
    grand_paid = grand_received = grand_approved = grand_rejected = 0.0

    for month_ym, group in df.groupby("month", sort=False):
        label = month_label(month_ym)

        # Month header row
        for ci in range(1, num_cols + 1):
            cell = ws.cell(row=current_row, column=ci, value=(label if ci == 1 else ""))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        current_row += 1

        # Data rows
        for row_idx, (_, row) in enumerate(group.iterrows()):
            row_fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
            for ci, (key, fmt) in enumerate(zip(_DETAIL_KEYS, _DETAIL_FMTS), 1):
                val = row.get(key)
                if not isinstance(val, str) and pd.isna(val):
                    val = None
                cell = ws.cell(row=current_row, column=ci, value=val)
                cell.fill = row_fill
                if fmt:
                    cell.number_format = fmt
            current_row += 1

        # Subtotal row
        m_paid     = float(group["paid"].sum())
        m_received = float(group["received"].sum())
        m_approved = float(group["approved"].sum())
        m_rejected = float(group["rejected"].sum())
        grand_paid     += m_paid
        grand_received += m_received
        grand_approved += m_approved
        grand_rejected += m_rejected

        subtotal_vals = {1: f"Subtotal — {label}", 5: m_paid, 6: m_received, 7: m_approved, 8: m_rejected}
        for ci in range(1, num_cols + 1):
            cell = ws.cell(row=current_row, column=ci, value=subtotal_vals.get(ci, ""))
            cell.fill = SUBTOTAL_FILL
            cell.font = SUBTOTAL_FONT
            if ci in (5, 6, 7, 8):
                cell.number_format = RUPEE_FMT
        current_row += 1

        # Blank spacer
        current_row += 1

    # Grand total row
    gt_vals = {1: "GRAND TOTAL", 5: grand_paid, 6: grand_received, 7: grand_approved, 8: grand_rejected}
    for ci in range(1, num_cols + 1):
        cell = ws.cell(row=current_row, column=ci, value=gt_vals.get(ci, ""))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        if ci in (5, 6, 7, 8):
            cell.number_format = RUPEE_FMT

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_fy_detail_report(df: pd.DataFrame, fy_label: str) -> bytes:
    """Generates a structured Excel workbook for the FY Admission Detail report."""
    return _generate_detail_report(df, f"FY {fy_label}")


def generate_month_detail_report(df: pd.DataFrame, title: str) -> bytes:
    """Generates a structured Excel workbook for the Month Admission Detail report."""
    return _generate_detail_report(df, title)


def generate_report(df: pd.DataFrame, title: str, report_type: str) -> bytes:
    """
    Generate an in-memory Excel workbook and return its bytes.

    report_type: 'admission_report' | 'monthly_summary' | 'fy_summary' | 'raw_export'
    """
    wb = Workbook()
    ws = wb.active

    if report_type == "admission_report":
        _write_sheet(ws, df, ADMISSION_COLS, title, status_col_idx=12)

    elif report_type == "monthly_summary":
        _write_sheet(ws, df, MONTHLY_COLS, title)

    elif report_type == "fy_summary":
        _write_sheet(ws, df, FY_COLS, title)

    elif report_type == "raw_export":
        # Use all columns from the DataFrame as-is
        col_defs = []
        for col in df.columns:
            fmt = RUPEE_FMT if col in AMOUNT_COLS else None
            col_defs.append((col.replace("_", " ").title(), col, fmt))
        _write_sheet(ws, df, col_defs, title)

    else:
        raise ValueError(f"Unknown report_type: {report_type}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _prepare_doctor_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy().reset_index(drop=True)
    out["no"] = range(1, len(out) + 1)
    # NULL and 0 both map to "No" — intentional, NULL means payment not yet recorded
    out["doctor_paid_label"] = out["doctor_paid"].apply(lambda x: "Yes" if x else "No")
    return out


_SUBTOTAL_FILL   = StylePatternFill(fill_type="solid", fgColor="FFF2CC")
_REJECTED_FILL   = StylePatternFill(fill_type="solid", fgColor="FCE4D6")
_SECTION_HDR_FONT = Font(bold=True, color="1F3864", name="Calibri")

# Column index within DOCTOR_INTERNAL_COLS (1-based) → field name
_DOCTOR_NUMERIC_COLS = {4: "hosp_ex", 5: "pharma_ex", 6: "dialysis_ex",
                        7: "total_ex", 8: "maa_payment", 9: "doctor_share",
                        10: "hospital_share"}


def _write_doctor_status_totals(ws, df: pd.DataFrame, n_cols: int | None = None) -> None:
    """Append status-wise subtotals for doctor-not-yet-paid entries."""
    unpaid = df[df["doctor_paid"] == 0].copy()
    if unpaid.empty:
        return

    if n_cols is None:
        n_cols = len(DOCTOR_INTERNAL_COLS)

    start_row = ws.max_row + 2

    hdr = ws.cell(row=start_row, column=1,
                  value="Outstanding Summary — Doctor Not Yet Paid")
    hdr.font = _SECTION_HDR_FONT
    start_row += 1

    sections = [
        ("Claim Paid",   unpaid[unpaid["maa_status"] == "Claim Paid"],  _SUBTOTAL_FILL),
        ("Rejected",     unpaid[unpaid["maa_status"] == "Rejected"],     _REJECTED_FILL),
        ("Total Unpaid", unpaid,                                         TOTAL_FILL),
    ]

    for label, sub, fill in sections:
        if sub.empty:
            continue
        ws.cell(row=start_row, column=1, value=label).font = TOTAL_FONT
        ws.cell(row=start_row, column=2,
                value=f"({len(sub)} {'entry' if len(sub) == 1 else 'entries'})"
                ).font = Font(italic=True, name="Calibri")
        for col_idx, field in _DOCTOR_NUMERIC_COLS.items():
            if field in sub.columns:
                val = sub[field].fillna(0).sum()
                cell = ws.cell(row=start_row, column=col_idx, value=val)
                cell.number_format = RUPEE_FMT
                cell.font = TOTAL_FONT
        for ci in range(1, n_cols + 1):
            ws.cell(row=start_row, column=ci).fill = fill
        start_row += 1


def generate_doctor_internal(df: pd.DataFrame, month_label: str) -> bytes:
    """Full internal report: all columns including payment tracking.

    month_label should be short (e.g. "June 2025") — sheet titles are truncated to 31 chars.
    """
    wb = Workbook()
    ws = wb.active
    _write_sheet(ws, _prepare_doctor_df(df), DOCTOR_INTERNAL_COLS, f"{month_label} (Internal)",
                 color_by_maa_status=True)
    _write_doctor_status_totals(ws, df, n_cols=len(DOCTOR_INTERNAL_COLS))
    _auto_width(ws)
    ws.column_dimensions['A'].width = min(ws.column_dimensions['A'].width, 20)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_doctor_copy(df: pd.DataFrame, month_label: str) -> bytes:
    """Doctor-facing report: no payment tracking columns."""
    wb = Workbook()
    ws = wb.active
    _write_sheet(
        ws, _prepare_doctor_df(df), DOCTOR_COPY_COLS, f"{month_label} (Dr Copy)",
        color_by_maa_status=True,
        header_fill=DOCTOR_COPY_HEADER_FILL, header_font=DOCTOR_COPY_HEADER_FONT,
    )
    _write_doctor_status_totals(ws, df, n_cols=len(DOCTOR_COPY_COLS))
    _auto_width(ws)
    ws.column_dimensions['A'].width = min(ws.column_dimensions['A'].width, 20)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
