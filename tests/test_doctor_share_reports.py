import io
import pandas as pd
import pytest
from openpyxl import load_workbook
import reports


@pytest.fixture
def sample_entries():
    return pd.DataFrame([
        {
            "id": 1, "tid": "T001", "patient_name": "Ravi Kumar",
            "admission_date": "2025-06-15", "month": "2025-06",
            "hosp_ex": 500.0, "pharma_ex": 1000.0, "dialysis_ex": 0.0,
            "total_ex": 1500.0, "maa_payment": 24300.0,
            "doctor_share": 9120.0, "hospital_share": 13680.0,
            "doctor_pct": 0.4, "doctor_flat": None,
            "maa_status": "Claim Paid", "doctor_paid": 1,
            "doctor_payment_month": "2025-06", "comments": None,
        },
        {
            "id": 2, "tid": None, "patient_name": "Cash Patient",
            "admission_date": "2025-06-20", "month": "2025-06",
            "hosp_ex": 0.0, "pharma_ex": 0.0, "dialysis_ex": 0.0,
            "total_ex": 0.0, "maa_payment": None,
            "doctor_share": 3000.0, "hospital_share": None,
            "doctor_pct": 0.4, "doctor_flat": 3000.0,
            "maa_status": None, "doctor_paid": 0,
            "doctor_payment_month": None, "comments": "Non-MAA",
        },
    ])


def test_generate_doctor_internal_returns_bytes(sample_entries):
    result = reports.generate_doctor_internal(sample_entries, "June 2025")
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_generate_doctor_internal_sheet_name(sample_entries):
    wb = load_workbook(io.BytesIO(reports.generate_doctor_internal(sample_entries, "June 2025")))
    assert "June 2025 (Internal)" in wb.sheetnames


def test_generate_doctor_internal_column_headers(sample_entries):
    wb = load_workbook(io.BytesIO(reports.generate_doctor_internal(sample_entries, "June 2025")))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for expected in ["No.", "Patient Name", "Doctor Share", "Doctor Paid", "Dr Payment Month"]:
        assert expected in headers


def test_generate_doctor_copy_omits_payment_tracking(sample_entries):
    wb = load_workbook(io.BytesIO(reports.generate_doctor_copy(sample_entries, "June 2025")))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "Doctor Paid" not in headers
    assert "Dr Payment Month" not in headers


def test_generate_doctor_copy_sheet_name(sample_entries):
    wb = load_workbook(io.BytesIO(reports.generate_doctor_copy(sample_entries, "June 2025")))
    assert "June 2025 (Dr Copy)" in wb.sheetnames


def test_report_row_count(sample_entries):
    wb = load_workbook(io.BytesIO(reports.generate_doctor_internal(sample_entries, "June 2025")))
    ws = wb.active
    # header + 2 data rows + total row = rows 1-4
    # outstanding summary (entry 2, unpaid/Non-MAA): blank + header + "Non-MAA" row + "Total Outstanding" row = rows 6-8
    # payments summary (entry 1, paid in 2025-06): blank + header + "June 2025" row + "Total Paid" row = rows 10-12
    assert ws.max_row == 12


def test_generate_doctor_copy_header_is_teal(sample_entries):
    wb = load_workbook(io.BytesIO(reports.generate_doctor_copy(sample_entries, "June 2025")))
    ws = wb.active
    # openpyxl returns ARGB; last 6 chars are the hex colour
    header_fill = ws.cell(1, 1).fill.fgColor.rgb[-6:]
    assert header_fill == "00695C", f"Expected teal 00695C, got {header_fill}"


def test_unreconciled_hospital_share_is_flagged(sample_entries):
    df = sample_entries.copy()
    df.loc[0, "shares_reconcile"] = False
    df.loc[1, "shares_reconcile"] = True
    wb = load_workbook(io.BytesIO(reports.generate_doctor_internal(df, "June 2025")))
    ws = wb.active
    hosp_col = 9  # Hospital Share column in DOCTOR_INTERNAL_COLS
    flagged_cell = ws.cell(2, hosp_col)
    ok_cell = ws.cell(3, hosp_col)
    assert flagged_cell.font.color.rgb[-6:] == "C00000"
    assert flagged_cell.comment is not None
    assert ok_cell.comment is None


def test_generate_doctor_internal_header_is_navy(sample_entries):
    wb = load_workbook(io.BytesIO(reports.generate_doctor_internal(sample_entries, "June 2025")))
    ws = wb.active
    header_fill = ws.cell(1, 1).fill.fgColor.rgb[-6:]
    assert header_fill == "1F3864", f"Expected navy 1F3864, got {header_fill}"
